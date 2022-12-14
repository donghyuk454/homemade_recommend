import json
import openpyxl

small_breeds = ["스피츠", "시츄", "요크셔테리어", "말티즈", "닥스훈트", "포메라니안", "푸들", "치와와", "미니핀", "슈나우저", "페키니즈", "빠삐용", "잭러셀테리어 믹스", "장모치와와", "밀티푸", "실버 푸들"]

class Recipe:
    def __init__(self, json_data):
        self.name = json_data['name']
        self.age = json_data['age'] 
        self.disease = json_data['disease']
        self.favor = json_data['favor'] # score 를 위해 추가
        self.tools = json_data['tools']
        self.hard = json_data['hard']
        self.allergy = json_data['allergy']
    
    def get_name(self):
        return self.name

    def check(self, recipe_elememt, user_element):
        for r in recipe_elememt:
            if r in user_element:
                return True
        return False

    def filter(self, user):
        if self.check(self.disease, user.disease):
            print(self.name+"제외, 이유 : 질병")
            return False

        if self.check(self.tools, user.tools):
            print(self.name+"제외, 이유 : 집기")
            return False

        if self.check(self.allergy, user.allergy):
            print(self.name+"제외, 이유 : 알러지")
            return False
        
        if self.hard < user.hard[0] or self.hard > user.hard[1]:
            print(self.name+"제외, 이유 : 강도")
            return False

        return True


class User:
    hard = [1]
    def __init__(self, age, breed, disease, favor, tools, allergy):
        self.age = age
        hardEnd = 3
        if breed in small_breeds or age <= 12  or age >= 84:
            hardEnd = 2
        self.hard.append(hardEnd)
        self.disease = disease
        self.favor = favor
        self.tools = tools
        self.allergy = allergy

recipe_list = []

with open("./recipes.json", "r", encoding="UTF-8") as json_file:
    json_data = json.load(json_file)
    for data in json_data['recipes']:
        recipe = Recipe(json_data=data)
        recipe_list.append(recipe)

wb = openpyxl.load_workbook("./doginfo.xlsx", data_only=True)
sheet = wb['Sheet1']

users = []

for i in range(2, 94):
    breed = sheet.cell(row=i, column=2).value
    age = sheet.cell(row=i, column=3).value
    disease = []
    for j in range(4, 8):
        if sheet.cell(row=i, column=j) == 1:
            disease.append(j-3)
    favor = []
    for j in range(8, 13):
        if sheet.cell(row=i, column=j) == 1:
            favor.append(j-7)
    allergy = []
    for j in range(13, 19):
        if sheet.cell(row=i, column=j) == 1:
            allergy.append(j-12)
    tools = []
    for j in range(19, 24):
        if sheet.cell(row=i, column=j) == 1:
            tools.append(j-18)

    user = User(breed=breed, age=age, disease=disease, favor=favor, tools=tools, allergy=allergy)
    users.append(user)

results = []
user = users[0]

for recipe in recipe_list:
    results.append([recipe.get_name(), recipe.filter(user)])

print(results)